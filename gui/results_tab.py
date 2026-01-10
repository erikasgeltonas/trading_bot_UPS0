# gui/results_tab.py
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import csv

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None


class ResultsTab:
    """Backtest statistics tab (MT5-like table) + export."""

    def __init__(self, parent):
        self.frame = tk.Frame(parent)
        self._row_ids = {}
        self._build_layout()

    def _build_layout(self):
        root = self.frame

        container = tk.Frame(root, bg="#1e1e1e")
        container.pack(fill="both", expand=True)

        # Top bar
        topbar = tk.Frame(container, bg="#1e1e1e")
        topbar.grid(row=0, column=0, columnspan=2, sticky="ew")
        topbar.grid_columnconfigure(0, weight=1)

        ttk.Button(topbar, text="Export CSV", command=self.export_csv).pack(side="left", padx=8, pady=6)
        ttk.Button(topbar, text="Export XLSX", command=self.export_xlsx).pack(side="left", padx=8, pady=6)

        style = ttk.Style()
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure(
            "Results.Treeview",
            background="#1e1e1e",
            fieldbackground="#1e1e1e",
            foreground="#dcdcdc",
            rowheight=24,
            borderwidth=0,
        )
        style.configure(
            "Results.Treeview.Heading",
            background="#2a2a2a",
            foreground="#ffffff",
            relief="flat",
            padding=(6, 6),
        )
        style.map("Results.Treeview.Heading", background=[("active", "#333333")])

        tree_area = tk.Frame(container, bg="#1e1e1e")
        tree_area.grid(row=1, column=0, sticky="nsew")
        container.grid_rowconfigure(1, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.columns = ("all", "long", "short", "market")
        self.tree = ttk.Treeview(
            tree_area,
            columns=self.columns,
            show="tree headings",
            style="Results.Treeview",
            selectmode="browse",
        )

        self.tree.heading("#0", text="Metric")
        self.tree.heading("all", text="All")
        self.tree.heading("long", text="Long")
        self.tree.heading("short", text="Short")
        self.tree.heading("market", text="Market")

        self.tree.column("#0", width=320, anchor="w", stretch=True)
        for c in self.columns:
            self.tree.column(c, width=160, anchor="e", stretch=False)

        vsb = ttk.Scrollbar(tree_area, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_area, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        tree_area.grid_rowconfigure(0, weight=1)
        tree_area.grid_columnconfigure(0, weight=1)

        self.tree.tag_configure("section", background="#252525", foreground="#ffffff", font=("Arial", 10, "bold"))
        self.tree.tag_configure("pos", foreground="#4FC3F7")
        self.tree.tag_configure("neg", foreground="#FF8A65")
        self.tree.tag_configure("muted", foreground="#9e9e9e")

        self._build_rows()

    def _build_rows(self):
        self.tree.delete(*self.tree.get_children())
        self._row_ids.clear()

        def add_section(title: str):
            return self.tree.insert("", "end", text=title, values=("", "", "", ""), tags=("section",))

        def add_row(parent, label: str, key: str):
            rid = self.tree.insert(parent, "end", text=label, values=("-", "-", "-", "-"), tags=("muted",))
            self._row_ids[key] = rid
            return rid

        s = add_section("Summary")
        add_row(s, "File", "file")
        add_row(s, "Net profit", "net_profit")
        add_row(s, "Net profit %", "net_profit_pct")
        add_row(s, "Initial balance", "initial_balance")
        add_row(s, "Final balance", "final_balance")
        add_row(s, "Balance change %", "balance_change_pct")

        a = add_section("All trades")
        add_row(a, "Trades count", "trades")
        add_row(a, "Average PnL", "avg_pnl")
        add_row(a, "Win rate %", "win_rate")

        w = add_section("Winning trades")
        add_row(w, "Winning trades", "win_count")
        add_row(w, "Gross profit", "gross_profit")
        add_row(w, "Average win", "avg_win")
        add_row(w, "Max win", "max_win")

        l = add_section("Losing trades")
        add_row(l, "Losing trades", "loss_count")
        add_row(l, "Gross loss", "gross_loss")
        add_row(l, "Average loss", "avg_loss")
        add_row(l, "Max loss", "max_loss")

        d = add_section("Drawdowns")
        add_row(d, "Max drawdown", "max_drawdown")
        add_row(d, "Max drawdown %", "max_drawdown_pct")

        r = add_section("Ratios")
        add_row(r, "Profit factor", "profit_factor")
        add_row(r, "Recovery factor", "recovery_factor")
        add_row(r, "Payoff ratio (AvgWin/AvgLoss)", "payoff_ratio")
        add_row(r, "Loss recovery trades (|AvgLoss|/AvgWin)", "loss_recovery_trades")

        for item in self.tree.get_children():
            self.tree.item(item, open=True)

    @staticmethod
    def _fmt_number(x, decimals=2):
        try:
            return f"{float(x):.{decimals}f}"
        except Exception:
            return "-"

    @staticmethod
    def _fmt_pct(x, decimals=2):
        try:
            return f"{float(x):.{decimals}f} %"
        except Exception:
            return "-"

    def _set_row(self, key: str, all_value, long_value="-", short_value="-", market_value="-", force_tag=None):
        rid = self._row_ids.get(key)
        if not rid:
            return

        self.tree.item(rid, values=(all_value, long_value, short_value, market_value))

        if force_tag:
            self.tree.item(rid, tags=(force_tag,))
            return

        # tag by All only
        try:
            v = float(str(all_value).replace("%", "").strip())
            if v < 0:
                self.tree.item(rid, tags=("neg",))
            elif v > 0:
                self.tree.item(rid, tags=("pos",))
            else:
                self.tree.item(rid, tags=("muted",))
        except Exception:
            self.tree.item(rid, tags=("muted",))

    def update_results(self, meta: dict, stats_all: dict, stats_long: dict | None = None, stats_short: dict | None = None, stats_market: dict | None = None) -> None:
        stats_long = stats_long or {}
        stats_short = stats_short or {}
        stats_market = stats_market or {}

        history_path = meta.get("history_path") or "-"

        def getv(stats: dict, k: str, default=0.0):
            return stats.get(k, default)

        # --- Summary ---
        self._set_row("file", history_path, "-", "-", "-", force_tag="muted")

        initial_all = float(getv(stats_all, "initial_balance", meta.get("initial", 0.0) or 0.0) or 0.0)
        final_all = float(getv(stats_all, "final_balance", meta.get("final_balance", initial_all) or initial_all) or initial_all)
        pnl_all = float(getv(stats_all, "total_pnl", 0.0) or 0.0)

        def net_profit_pct(stats: dict):
            init = float(getv(stats, "initial_balance", initial_all) or 0.0)
            pnl = float(getv(stats, "total_pnl", 0.0) or 0.0)
            return (pnl / init * 100.0) if init else 0.0

        def balance_change_pct(stats: dict):
            init = float(getv(stats, "initial_balance", initial_all) or 0.0)
            fin = float(getv(stats, "final_balance", init) or init)
            return ((fin / init) - 1.0) * 100.0 if init else 0.0

        self._set_row(
            "net_profit",
            self._fmt_number(pnl_all),
            self._fmt_number(getv(stats_long, "total_pnl", "-")),
            self._fmt_number(getv(stats_short, "total_pnl", "-")),
            "-",
        )
        self._set_row(
            "net_profit_pct",
            self._fmt_pct(net_profit_pct(stats_all)),
            self._fmt_pct(net_profit_pct(stats_long)) if stats_long else "-",
            self._fmt_pct(net_profit_pct(stats_short)) if stats_short else "-",
            "-",
        )
        self._set_row(
            "initial_balance",
            self._fmt_number(initial_all),
            self._fmt_number(getv(stats_long, "initial_balance", "-")) if stats_long else "-",
            self._fmt_number(getv(stats_short, "initial_balance", "-")) if stats_short else "-",
            "-",
            force_tag="muted",
        )
        self._set_row(
            "final_balance",
            self._fmt_number(final_all),
            self._fmt_number(getv(stats_long, "final_balance", "-")) if stats_long else "-",
            self._fmt_number(getv(stats_short, "final_balance", "-")) if stats_short else "-",
            "-",
        )
        self._set_row(
            "balance_change_pct",
            self._fmt_pct(balance_change_pct(stats_all)),
            self._fmt_pct(balance_change_pct(stats_long)) if stats_long else "-",
            self._fmt_pct(balance_change_pct(stats_short)) if stats_short else "-",
            "-",
        )

        # --- Trades ---
        self._set_row(
            "trades",
            str(int(getv(stats_all, "trades", 0) or 0)),
            str(int(getv(stats_long, "trades", 0) or 0)) if stats_long else "-",
            str(int(getv(stats_short, "trades", 0) or 0)) if stats_short else "-",
            "-",
            force_tag="muted",
        )
        self._set_row(
            "avg_pnl",
            self._fmt_number(getv(stats_all, "avg_pnl", 0.0)),
            self._fmt_number(getv(stats_long, "avg_pnl", 0.0)) if stats_long else "-",
            self._fmt_number(getv(stats_short, "avg_pnl", 0.0)) if stats_short else "-",
            "-",
        )
        self._set_row(
            "win_rate",
            self._fmt_pct(getv(stats_all, "win_rate", 0.0)),
            self._fmt_pct(getv(stats_long, "win_rate", 0.0)) if stats_long else "-",
            self._fmt_pct(getv(stats_short, "win_rate", 0.0)) if stats_short else "-",
            "-",
        )

        # --- Win/Loss blocks ---
        self._set_row("win_count", str(int(getv(stats_all, "win_count", 0) or 0)),
                      str(int(getv(stats_long, "win_count", 0) or 0)) if stats_long else "-",
                      str(int(getv(stats_short, "win_count", 0) or 0)) if stats_short else "-",
                      "-", force_tag="muted")
        self._set_row("gross_profit", self._fmt_number(getv(stats_all, "gross_profit", 0.0)),
                      self._fmt_number(getv(stats_long, "gross_profit", 0.0)) if stats_long else "-",
                      self._fmt_number(getv(stats_short, "gross_profit", 0.0)) if stats_short else "-",
                      "-")
        self._set_row("avg_win", self._fmt_number(getv(stats_all, "avg_win", 0.0)),
                      self._fmt_number(getv(stats_long, "avg_win", 0.0)) if stats_long else "-",
                      self._fmt_number(getv(stats_short, "avg_win", 0.0)) if stats_short else "-",
                      "-")
        self._set_row("max_win", "-", "-", "-", "-", force_tag="muted")  # jei vėliau norėsi – paskaičiuosim

        self._set_row("loss_count", str(int(getv(stats_all, "loss_count", 0) or 0)),
                      str(int(getv(stats_long, "loss_count", 0) or 0)) if stats_long else "-",
                      str(int(getv(stats_short, "loss_count", 0) or 0)) if stats_short else "-",
                      "-", force_tag="muted")
        self._set_row("gross_loss", self._fmt_number(getv(stats_all, "gross_loss", 0.0)),
                      self._fmt_number(getv(stats_long, "gross_loss", 0.0)) if stats_long else "-",
                      self._fmt_number(getv(stats_short, "gross_loss", 0.0)) if stats_short else "-",
                      "-")
        self._set_row("avg_loss", self._fmt_number(getv(stats_all, "avg_loss", 0.0)),
                      self._fmt_number(getv(stats_long, "avg_loss", 0.0)) if stats_long else "-",
                      self._fmt_number(getv(stats_short, "avg_loss", 0.0)) if stats_short else "-",
                      "-")
        self._set_row("max_loss", "-", "-", "-", "-", force_tag="muted")  # jei vėliau norėsi – paskaičiuosim

        # --- Drawdown ---
        def dd_pct(stats: dict):
            init = float(getv(stats, "initial_balance", initial_all) or 0.0)
            dd = float(getv(stats, "max_drawdown", 0.0) or 0.0)
            return (dd / init * 100.0) if init else 0.0

        self._set_row("max_drawdown", self._fmt_number(getv(stats_all, "max_drawdown", 0.0)),
                      self._fmt_number(getv(stats_long, "max_drawdown", 0.0)) if stats_long else "-",
                      self._fmt_number(getv(stats_short, "max_drawdown", 0.0)) if stats_short else "-",
                      "-")
        self._set_row("max_drawdown_pct", self._fmt_pct(dd_pct(stats_all)),
                      self._fmt_pct(dd_pct(stats_long)) if stats_long else "-",
                      self._fmt_pct(dd_pct(stats_short)) if stats_short else "-",
                      "-")

        # --- Ratios ---
        self._set_row("profit_factor", self._fmt_number(getv(stats_all, "profit_factor", 0.0)),
                      self._fmt_number(getv(stats_long, "profit_factor", 0.0)) if stats_long else "-",
                      self._fmt_number(getv(stats_short, "profit_factor", 0.0)) if stats_short else "-",
                      "-")
        self._set_row("recovery_factor", self._fmt_number(getv(stats_all, "recovery_factor", 0.0)),
                      self._fmt_number(getv(stats_long, "recovery_factor", 0.0)) if stats_long else "-",
                      self._fmt_number(getv(stats_short, "recovery_factor", 0.0)) if stats_short else "-",
                      "-")
        self._set_row("payoff_ratio", self._fmt_number(getv(stats_all, "payoff_ratio", 0.0)),
                      self._fmt_number(getv(stats_long, "payoff_ratio", 0.0)) if stats_long else "-",
                      self._fmt_number(getv(stats_short, "payoff_ratio", 0.0)) if stats_short else "-",
                      "-")
        self._set_row("loss_recovery_trades", self._fmt_number(getv(stats_all, "loss_recovery_trades", 0.0)),
                      self._fmt_number(getv(stats_long, "loss_recovery_trades", 0.0)) if stats_long else "-",
                      self._fmt_number(getv(stats_short, "loss_recovery_trades", 0.0)) if stats_short else "-",
                      "-")

    # ---------- Export ----------
    def _iter_rows_for_export(self):
        def walk(parent_id="", level=0):
            for iid in self.tree.get_children(parent_id):
                metric = self.tree.item(iid, "text")
                vals = self.tree.item(iid, "values") or ("", "", "", "")
                vals = tuple(vals) + ("", "", "", "")
                yield (level, metric, vals[0], vals[1], vals[2], vals[3])
                yield from walk(iid, level + 1)

        yield from walk("", 0)

    def export_csv(self):
        path = filedialog.asksaveasfilename(
            title="Export Results to CSV",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
        )
        if not path:
            return

        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["Level", "Metric", "All", "Long", "Short", "Market"])
                for row in self._iter_rows_for_export():
                    w.writerow(row)
            messagebox.showinfo("Export", f"CSV exported:\n{path}")
        except Exception as e:
            messagebox.showerror("Export error", str(e))

    def export_xlsx(self):
        if Workbook is None:
            messagebox.showerror("Export error", "openpyxl is not available. Install openpyxl or use CSV export.")
            return

        path = filedialog.asksaveasfilename(
            title="Export Results to XLSX",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"

            ws.append(["Level", "Metric", "All", "Long", "Short", "Market"])
            for row in self._iter_rows_for_export():
                ws.append(list(row))

            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 42
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 18
            ws.column_dimensions["E"].width = 18
            ws.column_dimensions["F"].width = 18

            wb.save(path)
            messagebox.showinfo("Export", f"XLSX exported:\n{path}")
        except Exception as e:
            messagebox.showerror("Export error", str(e))
